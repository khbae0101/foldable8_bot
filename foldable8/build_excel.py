# -*- coding: utf-8 -*-
"""집계 + 엑셀 생성. 요약 시트(공지용) + raw 시트(CRM 포함, 메일용)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
QUAL_KEYS = ["120K", "110K", "2nd", "삼디초", "가전구독", "제휴카드", "라이프", "MIT"]
NUM_KEYS = (["예약당일", "예약누적", "폴드8당일", "폴드8누적", "와이드당일", "와이드누적",
             "플립8당일", "플립8누적", "모두의행복", "소규모법인",
             "CRM모수", "컨택완료", "컨택성공", "컨택보류", "컨택실패"] + QUAL_KEYS)
 
 
def aggregate(reports, stores_cfg, prev_close=None, prev_reports=None):
    """reports: {조직명: 파싱결과}, prev_close: {조직명: 전일 누적},
    prev_reports: {조직명: 전일 파싱결과} — 당일 미제출 매장은 전일 보고값 이월."""
    prev_close = prev_close or {}
    prev_reports = prev_reports or {}
    stores = []
    for s in stores_cfg["매장"]:
        rec = {"상권": s["상권"], "코드": s["코드"], "조직": s["조직"], "목표": s["목표"]}
        rpt = reports.get(s["조직"])
        carried = False
        if not rpt and s["조직"] in prev_reports:
            rpt = prev_reports[s["조직"]]
            carried = True
        if rpt:
            for k in NUM_KEYS:
                rec[k] = rpt.get(k, 0)
            if carried:
                rec["예약당일"] = 0  # 이월분은 당일 실적 없음
            rec["미제출"] = s["조직"] not in reports
            rec["이월"] = carried
            rec["데이터있음"] = True
            pv = prev_close.get(s["조직"])
            rec["전일마감"] = pv
            rec["증분"] = rec["예약누적"] - pv if pv is not None else rpt.get("예약당일", 0)
        else:
            for k in NUM_KEYS:
                rec[k] = 0
            rec["미제출"] = True
            rec["이월"] = False
            rec["데이터있음"] = False
            rec["전일마감"] = prev_close.get(s["조직"])
            rec["증분"] = None
        stores.append(rec)
 
    with_data = [s for s in stores if s["데이터있음"]]
    with_data.sort(key=lambda s: (s["예약누적"] / s["목표"]) if s["목표"] else 0,
                   reverse=True)
    for i, s in enumerate(with_data, 1):
        s["순위"] = i
 
    # 상/하위 25% 하이라이트 (달성률·순위 + 유치율 컬럼별)
    n = len(with_data)
    q = max(1, int(n * 0.25 + 0.5))
    for i, s in enumerate(with_data):
        s["달성률HL"] = "good" if i < q else ("bad" if i >= n - q else None)
    for key in QUAL_KEYS:
        ranked = sorted(with_data,
                        key=lambda s: (s[key] / s["예약누적"]) if s["예약누적"] else 0,
                        reverse=True)
        for i, s in enumerate(ranked):
            s.setdefault("유치율HL", {})[key] = \
                "good" if i < q else ("bad" if i >= n - q else None)
 
    def total(rows, name):
        t = {"조직": name, "코드": "", "상권": "", "목표": sum(r["목표"] for r in rows)}
        for k in NUM_KEYS:
            t[k] = sum(r[k] for r in rows)
        pvs = [r["전일마감"] for r in rows if r["전일마감"] is not None]
        t["전일마감"] = sum(pvs) if pvs else None
        t["증분"] = t["예약누적"] - t["전일마감"] if t["전일마감"] is not None else \
            sum(r["증분"] or 0 for r in rows)
        return t
 
    regions_order = []
    for s in stores_cfg["매장"]:
        if s["상권"] not in regions_order:
            regions_order.append(s["상권"])
    region_totals = {reg: total([r for r in stores if r["상권"] == reg], reg)
                     for reg in regions_order}
    ach = sorted(regions_order,
                 key=lambda g: region_totals[g]["예약누적"] / region_totals[g]["목표"],
                 reverse=True)
    for i, g in enumerate(ach, 1):
        region_totals[g]["순위"] = i
 
    return {
        "지사계": total(stores, "지사 계"),
        "상권": region_totals,
        "상권순서": regions_order,
        "매장": stores,
        "매장_정렬": with_data + [s for s in stores if not s["데이터있음"]],
    }
 
 
# ---------------------------------------------------------------- Excel ----
THIN = Side(style="thin", color="A0A0A0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
 
 
def _style(ws, cell, v, bold=False, fill=None, fmt=None, color=None):
    c = ws[cell]
    c.value = v
    c.font = Font(name="맑은 고딕", size=10, bold=bold,
                  color=color or "000000")
    c.alignment = CENTER
    c.border = BORDER
    if fill:
        c.fill = PatternFill("solid", start_color=fill)
    if fmt:
        c.number_format = fmt
    return c
 
 
def build_workbook(agg, date_str, out_path, title="■ 강동소매 폴더블8 매장별 예약현황"):
    wb = Workbook()
 
    # ---------- 요약 시트 ----------
    ws = wb.active
    ws.title = "요약"
    ws.sheet_view.showGridLines = False
 
    heads = [("상권", 1), ("코드", 1), ("조직", 1), ("목표", 1), ("예약", 1), ("증분", 1),
             ("달성률", 1), ("순위", 1), ("모델별", 3), ("모두의\n행복", 1), ("소규모\n/법인", 1)]
    heads += [(k if k != "삼디초" else "삼/디초", 2) for k in QUAL_KEYS]
    heads += [("전일마감", 1)]
 
    ws.merge_cells("B2:H2")
    _style(ws, "B2", title, bold=True, fill="3F3F3F", color="FFFFFF")
    ws["B2"].font = Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF")
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    n_cols = sum(n for _, n in heads)
    last_col = get_column_letter(1 + n_cols)
    _style(ws, f"{last_col}2", date_str, bold=True, color="C00000")
    ws.row_dimensions[2].height = 28
 
    r0 = 4
    col = 2
    for name, span in heads:
        c1 = get_column_letter(col)
        c2 = get_column_letter(col + span - 1)
        fill = "FFF200" if name == "전일마감" else "D9D9D9"
        _style(ws, f"{c1}{r0}", name, bold=True, fill=fill)
        if span > 1:
            subs = ["폴드8", "와이드", "플립8"] if name == "모델별" else ["건", "유치율"]
            for i, sub in enumerate(subs):
                _style(ws, f"{get_column_letter(col + i)}{r0}", name if i == 0 else None,
                       bold=True, fill=fill) if i else None
                _style(ws, f"{get_column_letter(col + i)}{r0 + 1}", sub,
                       bold=True, fill=fill)
            ws.merge_cells(f"{c1}{r0}:{c2}{r0}")
        else:
            _style(ws, f"{c1}{r0 + 1}", None, fill=fill)
            ws.merge_cells(f"{c1}{r0}:{c1}{r0 + 1}")
        col += span
 
    HL_FILL = {"good": "C6EFCE", "bad": "FFC7CE"}
    HL_FONT = {"good": "006100", "bad": "9C0006"}
 
    def write_row(r, rec, kind):
        fills = {"total": "262626", "region": "E4DFEC"}
        fill = fills.get(kind)
        color = "FFFFFF" if kind == "total" else None
        bold = kind in ("total", "region")
        g = rec["예약누적"]
        name = rec["조직"] + ("＊" if rec.get("이월") else "")
        cells = [rec["상권"], rec["코드"], name, rec["목표"], rec["예약누적"],
                 rec["증분"],
                 rec["예약누적"] / rec["목표"] if rec["목표"] else None,
                 rec.get("순위"),
                 rec["폴드8누적"], rec["와이드누적"], rec["플립8누적"],
                 rec["모두의행복"], rec["소규모법인"]]
        for k in QUAL_KEYS:
            cells += [rec[k], rec[k] / g if g else None]
        cells += [rec["전일마감"]]
        no_data = not rec.get("데이터있음", True)
        if no_data:
            cells = cells[:4] + ["미제출"] + [None] * (len(cells) - 5)
        hl = {}
        if kind == "store" and not no_data:
            a = rec.get("달성률HL")
            if a:
                hl[6] = hl[7] = a
            for idx, k in enumerate(QUAL_KEYS):
                h = rec.get("유치율HL", {}).get(k)
                if h:
                    hl[14 + 2 * idx] = h
        for i, v in enumerate(cells):
            cl = get_column_letter(2 + i)
            fmt = None
            if cl_is_rate(i):
                fmt = "0.0%" if i == 6 else "0%"
            cell_fill = fill
            cell_color = color
            if kind == "store" and i == 3:
                cell_fill = "E2EFDA"
            if kind.startswith("store") and i == len(cells) - 1:
                cell_fill = "FFF2A8"
            if i in hl:
                cell_fill = HL_FILL[hl[i]]
                cell_color = HL_FONT[hl[i]]
            if no_data:
                cell_color = "969696"
            _style(ws, f"{cl}{r}", v, bold=bold, fill=cell_fill, fmt=fmt,
                   color=cell_color)
 
    def cl_is_rate(i):
        if i == 6:  # 달성률
            return True
        return i >= 13 and (i - 13) % 2 == 1 and i < 13 + 16
 
    r = r0 + 2
    write_row(r, agg["지사계"], "total"); r += 1
    for reg in agg["상권순서"]:
        write_row(r, agg["상권"][reg], "region"); r += 1
    for s in agg["매장_정렬"]:
        write_row(r, s, "store"); r += 1
    carried = [s["조직"] for s in agg["매장_정렬"] if s.get("이월")]
    nodata = [s["조직"] for s in agg["매장_정렬"] if not s.get("데이터있음")]
    if carried or nodata:
        note = []
        if carried:
            note.append(f"＊ 미제출 매장(전일값 이월): {', '.join(carried)}")
        if nodata:
            note.append(f"미제출(이월 데이터 없음): {', '.join(nodata)}")
        _style(ws, f"B{r + 1}", " / ".join(note), bold=True, color="C00000")
        ws[f"B{r + 1}"].alignment = Alignment(horizontal="left")
        ws[f"B{r + 1}"].border = Border()
 
    widths = [9, 10, 12, 6.5, 6.5, 6, 8, 5.5, 6.5, 6.5, 6.5, 7, 7] + [6, 7] * 8 + [8.5]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(2 + i)].width = w
 
    # ---------- raw 시트 (CRM 포함, 메일용) ----------
    ws2 = wb.create_sheet("raw(CRM포함)")
    cols2 = ["상권", "코드", "조직", "목표",
             "CRM 모수", "컨택완료", "컨택 성공", "컨택 보류", "컨택 실패", "컨택률",
             "예약 당일", "예약 누적", "달성률", "증분",
             "폴드8 당일", "폴드8 누적", "와이드 당일", "와이드 누적",
             "플립8 당일", "플립8 누적", "모두의 행복", "소규모/법인 특판"] + \
            [k if k != "삼디초" else "삼/디초" for k in QUAL_KEYS] + ["전일마감", "제출여부"]
    for i, h in enumerate(cols2):
        _style(ws2, f"{get_column_letter(1 + i)}1", h, bold=True, fill="D9D9D9")
        ws2.column_dimensions[get_column_letter(1 + i)].width = max(8, len(h) * 1.6 + 3)
    r = 2
    rows_all = [("total", agg["지사계"])] + \
               [("region", agg["상권"][g]) for g in agg["상권순서"]] + \
               [("store", s) for s in agg["매장_정렬"]]
    for kind, rec in rows_all:
        vals = [rec["상권"], rec["코드"], rec["조직"], rec["목표"],
                rec["CRM모수"], rec["컨택완료"], rec["컨택성공"], rec["컨택보류"],
                rec["컨택실패"],
                rec["컨택완료"] / rec["CRM모수"] if rec["CRM모수"] else None,
                rec["예약당일"], rec["예약누적"],
                rec["예약누적"] / rec["목표"] if rec["목표"] else None,
                rec["증분"],
                rec["폴드8당일"], rec["폴드8누적"], rec["와이드당일"], rec["와이드누적"],
                rec["플립8당일"], rec["플립8누적"], rec["모두의행복"], rec["소규모법인"]]
        vals += [rec[k] for k in QUAL_KEYS]
        status = "제출"
        if rec.get("이월"):
            status = "미제출(전일이월)"
        elif rec.get("미제출"):
            status = "미제출"
        vals += [rec["전일마감"], status]
        fill = {"total": "262626", "region": "E4DFEC"}.get(kind)
        color = "FFFFFF" if kind == "total" else None
        for i, v in enumerate(vals):
            fmt = "0.0%" if i in (9, 12) else None
            _style(ws2, f"{get_column_letter(1 + i)}{r}", v,
                   bold=kind != "store", fill=fill, fmt=fmt, color=color)
        r += 1
    ws2.freeze_panes = "D2"
 
    wb.save(out_path)
    return out_path
 
